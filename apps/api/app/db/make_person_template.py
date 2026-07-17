"""生成人员导入 Excel 模板（方案 3.1）。运行：python -m app.db.make_person_template <输出路径>"""
from __future__ import annotations

import sys

from openpyxl import Workbook
from openpyxl.styles import Font

HEADERS = ["学号", "班级", "姓名", "手机号", "困难等级", "身份证号", "银行卡号"]
SAMPLE = ["20230001", "计算机1班", "王文博", "13800001111", "一般", "", ""]


def build(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "人员导入"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(SAMPLE)
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    wb.save(path)
    print(f"已生成模板：{path}")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "人员导入模板.xlsx"
    build(out)
