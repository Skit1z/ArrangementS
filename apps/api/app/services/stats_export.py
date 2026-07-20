"""月度工时 Excel 导出（方案 11.3）。

两个工作表：月度汇总、工时明细。分场地列按场地维度动态生成。
身份证号、银行卡号绝不进入导出。
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.enums import SlotSourceType
from app.models.person import PersonProfile
from app.models.schedule import Assignment, DutySlot
from app.models.statistics import MonthlyHourSummary, MonthlyVenueHourSummary
from app.models.venue import Venue
from app.services.schedule_stats import month_to_date


def _h(minutes: int) -> float:
    return round(minutes / 60, 2)


def build_export(db: Session, month_key: str) -> bytes:
    month = month_to_date(month_key)
    wb = Workbook()

    _summary_sheet(db, wb, month_key, month)
    _detail_sheet(db, wb, month_key)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _summary_sheet(db: Session, wb: Workbook, month_key: str, month) -> None:
    ws = wb.active
    ws.title = "月度汇总"

    summaries = list(
        db.scalars(select(MonthlyHourSummary).where(MonthlyHourSummary.month == month))
    )
    venue_rows = list(
        db.scalars(select(MonthlyVenueHourSummary).where(MonthlyVenueHourSummary.month == month))
    )
    venue_ids = sorted({v.venue_id for v in venue_rows}, key=str)
    venue_names = (
        {v.id: v.name for v in db.scalars(select(Venue).where(Venue.id.in_(venue_ids)))}
        if venue_ids
        else {}
    )
    venue_completed = {(v.person_id, v.venue_id): v.completed_minutes for v in venue_rows}

    persons = (
        {
            p.id: p
            for p in db.scalars(
                select(PersonProfile).where(PersonProfile.id.in_([s.person_id for s in summaries]))
            )
        }
        if summaries
        else {}
    )

    headers = ["学号", "班级", "姓名", "手机号", "排班平衡工时(h)", "实际完成工时(h)"]
    headers += [f"{venue_names.get(vid, vid)}工时(h)" for vid in venue_ids]
    headers += ["倍率工时(h)", "请假次数", "换班转出次数", "未到岗次数"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for s in summaries:
        p = persons.get(s.person_id)
        row = [
            p.student_no if p else "",
            p.class_name if p else "",
            p.full_name if p else "",
            p.phone if p else "",
            _h(s.balance_minutes),
            _h(s.completed_minutes),
        ]
        row += [_h(venue_completed.get((s.person_id, vid), 0)) for vid in venue_ids]
        row += [_h(s.multiplier_extra_minutes), s.leave_count, s.swap_out_count, s.absence_count]
        ws.append(row)


def _detail_sheet(db: Session, wb: Workbook, month_key: str) -> None:
    ws = wb.create_sheet("工时明细")
    headers = [
        "日期",
        "场地",
        "班次/任务",
        "完整值班开始",
        "完整值班结束",
        "原始时长(min)",
        "取整前加权(min)",
        "最终统计工时(min)",
        "执行状态",
        "分配来源",
        "备注",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rows = db.execute(
        select(Assignment, DutySlot, Venue, PersonProfile)
        .join(DutySlot, Assignment.duty_slot_id == DutySlot.id)
        .join(Venue, DutySlot.venue_id == Venue.id)
        .join(PersonProfile, Assignment.person_id == PersonProfile.id)
        .where(DutySlot.month_key == month_key)
        .order_by(DutySlot.slot_start_at)
    ).all()
    for a, slot, venue, person in rows:
        kind = "黄楼班次" if slot.source_type == SlotSourceType.fixed_shift else "场地任务"
        ws.append(
            [
                slot.slot_start_at.date().isoformat(),
                venue.name,
                f"{kind}",
                slot.slot_start_at.isoformat(),
                slot.slot_end_at.isoformat(),
                a.raw_minutes,
                float(a.weighted_minutes_before_round),
                a.credited_minutes,
                a.execution_status.value,
                a.assignment_source.value,
                person.full_name,
            ]
        )
