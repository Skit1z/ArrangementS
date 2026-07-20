"""人员 Excel 导入（方案 3.1 / 3.2）。

流程：上传 -> 校验表头 -> 逐行校验 -> 预览(新增/更新/无变化/错误) -> admin 确认 -> 事务写入
      -> 自动生成/更新账号 -> 记录批次。

安全：敏感字段（身份证号/银行卡号）在预览阶段即加密，preview_payload 仅存密文(base64)+后四位，
明文绝不落库；初始密码可由“学号+拼音首字母”确定性重算，不存储、不入日志。
"""

from __future__ import annotations

import base64
import hashlib
import io
import re
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.crypto import encrypt_field, last4
from app.core.pinyin import build_initial_password
from app.core.security import hash_password
from app.models.enums import ImportBatchStatus, PersonStatus, UserRole
from app.models.import_batch import ImportBatch
from app.models.person import PersonProfile
from app.models.user import User

REQUIRED_COLUMNS = ["学号", "班级", "姓名", "手机号", "困难等级"]
OPTIONAL_COLUMNS = ["身份证号", "银行卡号"]

_PHONE_RE = re.compile(r"^1[3-9]\d{9}$")
_ID_CARD_RE = re.compile(r"^\d{17}[\dXx]$")
_BANK_CARD_RE = re.compile(r"^\d{12,19}$")

ROW_NEW = "new"
ROW_UPDATE = "update"
ROW_UNCHANGED = "unchanged"
ROW_ERROR = "error"


@dataclass
class RowResult:
    row_no: int
    student_no: str = ""
    class_name: str = ""
    full_name: str = ""
    phone: str = ""
    difficulty_level: str = ""
    id_card_last4: str | None = None
    bank_card_last4: str | None = None
    # 密文以 base64 存于预览负载，确认时直接写库；明文不落库
    _id_card_cipher_b64: str | None = None
    _bank_card_cipher_b64: str | None = None
    status: str = ROW_NEW
    errors: list[str] = field(default_factory=list)


def _cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_rows(file_bytes: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [_cell(c) for c in next(rows_iter)]
    except StopIteration:
        raise HTTPException(status_code=422, detail="Excel 为空")

    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        raise HTTPException(status_code=422, detail=f"缺少必填列：{'、'.join(missing)}")

    col_index = {name: header.index(name) for name in header if name}
    records: list[dict[str, str]] = []
    for raw in rows_iter:
        if raw is None or all(_cell(c) == "" for c in raw):
            continue
        rec = {}
        for name, idx in col_index.items():
            rec[name] = _cell(raw[idx]) if idx < len(raw) else ""
        records.append(rec)
    wb.close()
    return records


def _validate_row(row_no: int, rec: dict[str, str], seen: set[str]) -> RowResult:
    r = RowResult(row_no=row_no)
    r.student_no = rec.get("学号", "").strip()
    r.class_name = rec.get("班级", "").strip()
    r.full_name = rec.get("姓名", "").strip()
    r.phone = rec.get("手机号", "").strip()
    r.difficulty_level = rec.get("困难等级", "").strip()
    id_card = rec.get("身份证号", "").strip()
    bank_card = rec.get("银行卡号", "").strip()

    if not r.student_no:
        r.errors.append("学号为空")
    elif r.student_no in seen:
        r.errors.append("同批次学号重复")
    if not r.full_name:
        r.errors.append("姓名为空")
    if not r.class_name:
        r.errors.append("班级为空")
    if not _PHONE_RE.match(r.phone):
        r.errors.append("手机号格式非法")
    if id_card and not _ID_CARD_RE.match(id_card):
        r.errors.append("身份证号格式非法")
    if bank_card and not _BANK_CARD_RE.match(bank_card):
        r.errors.append("银行卡号格式非法")

    if r.student_no:
        seen.add(r.student_no)

    if id_card and not any("身份证号" in e for e in r.errors):
        r.id_card_last4 = last4(id_card)
        r._id_card_cipher_b64 = base64.b64encode(encrypt_field(id_card)).decode()
    if bank_card and not any("银行卡号" in e for e in r.errors):
        r.bank_card_last4 = last4(bank_card)
        r._bank_card_cipher_b64 = base64.b64encode(encrypt_field(bank_card)).decode()

    if r.errors:
        r.status = ROW_ERROR
    return r


def _classify(db: Session, rows: list[RowResult]) -> None:
    valid_nos = [r.student_no for r in rows if r.status != ROW_ERROR and r.student_no]
    existing = {
        p.student_no: p
        for p in db.scalars(select(PersonProfile).where(PersonProfile.student_no.in_(valid_nos)))
    }
    for r in rows:
        if r.status == ROW_ERROR:
            continue
        prof = existing.get(r.student_no)
        if prof is None:
            r.status = ROW_NEW
        elif (
            prof.class_name == r.class_name
            and prof.full_name == r.full_name
            and prof.phone == r.phone
            and (prof.difficulty_level or "") == r.difficulty_level
            and (prof.id_card_last4 or None) == r.id_card_last4
            and (prof.bank_card_last4 or None) == r.bank_card_last4
        ):
            r.status = ROW_UNCHANGED
        else:
            r.status = ROW_UPDATE


def _serialize(r: RowResult) -> dict:
    return asdict(r)


def create_preview(
    db: Session, file_name: str, file_bytes: bytes, actor_id: uuid.UUID
) -> ImportBatch:
    records = _read_rows(file_bytes)
    seen: set[str] = set()
    rows = [_validate_row(i + 2, rec, seen) for i, rec in enumerate(records)]
    _classify(db, rows)

    counts = {ROW_NEW: 0, ROW_UPDATE: 0, ROW_UNCHANGED: 0, ROW_ERROR: 0}
    for r in rows:
        counts[r.status] += 1

    batch = ImportBatch(
        file_name=file_name,
        file_hash=hashlib.sha256(file_bytes).hexdigest(),
        status=ImportBatchStatus.previewing,
        total_rows=len(rows),
        new_rows=counts[ROW_NEW],
        updated_rows=counts[ROW_UPDATE],
        error_rows=counts[ROW_ERROR],
        preview_payload={"rows": [_serialize(r) for r in rows], "counts": counts},
        created_by=actor_id,
    )
    db.add(batch)
    db.flush()
    return batch


@dataclass
class CreatedAccount:
    student_no: str
    full_name: str
    initial_password: str


def confirm_import(db: Session, batch_id: uuid.UUID) -> list[CreatedAccount]:
    """事务内确认写入。返回本批次新增账号的初始密码（一次性，供 admin 下载）。"""
    batch = db.get(ImportBatch, batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="导入批次不存在")
    if batch.status == ImportBatchStatus.confirmed:
        raise HTTPException(status_code=409, detail="该批次已确认")
    payload = batch.preview_payload or {}
    rows = payload.get("rows", [])

    created: list[CreatedAccount] = []
    for row in rows:
        if row["status"] in (ROW_ERROR, ROW_UNCHANGED):
            continue
        id_cipher = (
            base64.b64decode(row["_id_card_cipher_b64"]) if row.get("_id_card_cipher_b64") else None
        )
        bank_cipher = (
            base64.b64decode(row["_bank_card_cipher_b64"])
            if row.get("_bank_card_cipher_b64")
            else None
        )
        prof = db.scalar(select(PersonProfile).where(PersonProfile.student_no == row["student_no"]))
        if prof is None:
            initial = build_initial_password(row["student_no"], row["full_name"])
            user = User(
                username=row["student_no"],
                password_hash=hash_password(initial),
                role=UserRole.user,
                is_active=True,
            )
            db.add(user)
            db.flush()
            prof = PersonProfile(
                user_id=user.id,
                student_no=row["student_no"],
                class_name=row["class_name"],
                full_name=row["full_name"],
                phone=row["phone"],
                difficulty_level=row["difficulty_level"] or None,
                id_card_ciphertext=id_cipher,
                id_card_last4=row.get("id_card_last4"),
                bank_card_ciphertext=bank_cipher,
                bank_card_last4=row.get("bank_card_last4"),
                status=PersonStatus.active,
            )
            db.add(prof)
            created.append(CreatedAccount(row["student_no"], row["full_name"], initial))
        else:
            # 更新既有档案（不重置密码、不动 is_in_scheduling_pool 与账号启用状态）
            prof.class_name = row["class_name"]
            prof.full_name = row["full_name"]
            prof.phone = row["phone"]
            prof.difficulty_level = row["difficulty_level"] or None
            if id_cipher is not None:
                prof.id_card_ciphertext = id_cipher
                prof.id_card_last4 = row.get("id_card_last4")
            if bank_cipher is not None:
                prof.bank_card_ciphertext = bank_cipher
                prof.bank_card_last4 = row.get("bank_card_last4")

    batch.status = ImportBatchStatus.confirmed
    batch.confirmed_at = datetime.now(timezone.utc)
    # 清理预览负载中的密文，确认后不再需要
    batch.preview_payload = {"counts": payload.get("counts", {})}
    db.flush()
    return created
