"""课表编排：抽取结果 -> 课程规则(预览) -> 审核生效 -> 不可值班区间生成（方案 4.6 / 5.1）。

关键红线：
- 课表未经确认不得直接影响排班（写 CourseRule 但不生成 AvailabilityBlock，直至 approve）。
- 学期结束后旧课表与其生成的不可值班区间自动逻辑失效。
"""

from __future__ import annotations

import uuid
from datetime import datetime, timezone

from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.availability import AvailabilityBlock
from app.models.enums import (
    AvailabilitySource,
    AvailabilityStatus,
    ParseStatus,
    ReviewStatus,
)
from app.models.semester import CoursePeriodRule, Semester
from app.models.timetable import CourseRule, TimetableUpload
from app.services.intervals import merge_intervals
from app.services.semester_service import resolve_building_type
from app.timetable.availability import (
    generate_intervals,
    period_time_from_rule,
    resolve_period_time,
)
from app.timetable.extractor import RawCourseEntry
from app.timetable.week_parser import WeekParseError, parse_weeks

PARSER_VERSION = "v1"


def create_upload_from_entries(
    db: Session,
    *,
    person_id: uuid.UUID,
    semester_id: uuid.UUID,
    uploader_user_id: uuid.UUID | None,
    file_name: str,
    entries: list[RawCourseEntry],
    file_hash: str | None = None,
    storage_key: str | None = None,
) -> TimetableUpload:
    semester = db.get(Semester, semester_id)
    if semester is None:
        raise HTTPException(status_code=404, detail="学期不存在")

    upload = TimetableUpload(
        person_id=person_id,
        semester_id=semester_id,
        uploader_user_id=uploader_user_id,
        file_name=file_name,
        file_hash=file_hash,
        storage_key=storage_key,
        parse_status=ParseStatus.parsed,
        review_status=ReviewStatus.draft,
        parser_version=PARSER_VERSION,
    )
    db.add(upload)
    db.flush()

    orm_rules = db.scalars(
        select(CoursePeriodRule).where(
            CoursePeriodRule.semester_id == semester_id,
            CoursePeriodRule.is_active.is_(True),
        )
    )
    period_rules = [period_time_from_rule(r) for r in orm_rules]

    for entry in entries:
        upload.course_rules.append(
            _build_course_rule(db, semester_id, period_rules, entry, max_week=semester.week_count)
        )
    db.flush()
    return upload


def _build_course_rule(
    db: Session, semester_id: uuid.UUID, period_rules, entry: RawCourseEntry, *, max_week: int = 20
) -> CourseRule:
    needs_review = False
    warnings: list[str] = []

    building_type = None
    if entry.location_code:
        building_type = resolve_building_type(db, semester_id, entry.location_code)
    if building_type is None:
        needs_review = True
        warnings.append("教学楼代码未识别")

    week_start = week_end = None
    parity = "all"
    explicit_weeks: list[int] = []
    try:
        spec = parse_weeks(entry.week_expr, max_week=max_week)
        week_start, week_end, parity = spec.week_start, spec.week_end, spec.parity
        explicit_weeks = spec.explicit_weeks
        if spec.warnings:
            needs_review = True
    except WeekParseError:
        needs_review = True

    start_time = end_time = None
    if building_type is not None:
        resolved = resolve_period_time(
            period_rules, building_type, entry.period_start, entry.period_end
        )
        if resolved is None:
            needs_review = True
        else:
            start_time, end_time = resolved

    if entry.confidence is not None and entry.confidence < 0.8:
        needs_review = True

    return CourseRule(
        course_name=entry.course_name,
        weekday=entry.weekday,
        period_start=entry.period_start,
        period_end=entry.period_end,
        week_start=week_start,
        week_end=week_end,
        week_parity=parity,
        explicit_weeks=explicit_weeks or None,
        location_code=entry.location_code,
        building_type=building_type,
        start_time=start_time,
        end_time=end_time,
        confidence=entry.confidence,
        needs_review=needs_review,
    )


def _period_rules_for(db: Session, semester_id: uuid.UUID):
    orm_rules = db.scalars(
        select(CoursePeriodRule).where(
            CoursePeriodRule.semester_id == semester_id,
            CoursePeriodRule.is_active.is_(True),
        )
    )
    return [period_time_from_rule(r) for r in orm_rules]


def update_course_rule(
    db: Session, upload_id: uuid.UUID, rule_id: uuid.UUID, patch: dict
) -> CourseRule:
    """人工修正单条课程规则并重新解析（周次/建筑/时间/needs_review）。"""
    upload = _get_upload(db, upload_id)
    rule = db.get(CourseRule, rule_id)
    if rule is None or rule.timetable_upload_id != upload.id:
        raise HTTPException(status_code=404, detail="课程规则不存在")

    if "course_name" in patch and patch["course_name"] is not None:
        rule.course_name = patch["course_name"]
    if patch.get("weekday") is not None:
        rule.weekday = patch["weekday"]
    if patch.get("period_start") is not None:
        rule.period_start = patch["period_start"]
    if patch.get("period_end") is not None:
        rule.period_end = patch["period_end"]
    if patch.get("location_code") is not None:
        rule.location_code = patch["location_code"]

    needs_review = False
    rule.building_type = (
        resolve_building_type(db, upload.semester_id, rule.location_code)
        if rule.location_code
        else None
    )
    if rule.building_type is None:
        needs_review = True

    if patch.get("week_expr"):
        semester = db.get(Semester, upload.semester_id)
        max_week = semester.week_count if semester else 20
        try:
            spec = parse_weeks(patch["week_expr"], max_week=max_week)
            rule.week_start, rule.week_end = spec.week_start, spec.week_end
            rule.week_parity = spec.parity
            rule.explicit_weeks = spec.explicit_weeks
            if spec.warnings:
                needs_review = True
        except WeekParseError:
            needs_review = True

    if rule.building_type is not None:
        resolved = resolve_period_time(
            _period_rules_for(db, upload.semester_id),
            rule.building_type,
            rule.period_start,
            rule.period_end,
        )
        if resolved is None:
            needs_review = True
        else:
            rule.start_time, rule.end_time = resolved

    rule.needs_review = patch.get("needs_review", needs_review)
    db.flush()
    return rule


def submit_for_review(db: Session, upload_id: uuid.UUID) -> TimetableUpload:
    upload = _get_upload(db, upload_id)
    upload.review_status = ReviewStatus.pending
    db.flush()
    return upload


def approve(db: Session, upload_id: uuid.UUID, reviewer_id: uuid.UUID | None) -> TimetableUpload:
    """审核通过：使旧生效版本失效，生成本次不可值班区间。"""
    upload = _get_upload(db, upload_id)
    semester = db.get(Semester, upload.semester_id)

    _supersede_previous(db, upload)

    intervals: list[tuple[datetime, datetime]] = []
    for rule in upload.course_rules:
        if rule.start_time is None or rule.end_time is None or not rule.explicit_weeks:
            continue  # 未解析完整的规则不生成区间，留待人工修正后再审
        buffer = semester.course_buffer_minutes if semester.course_buffer_enabled else 0
        intervals.extend(
            generate_intervals(
                semester.first_monday,
                rule.weekday,
                rule.start_time,
                rule.end_time,
                rule.explicit_weeks,
                buffer,
            )
        )

    now = datetime.now(timezone.utc)
    for start, end in merge_intervals(intervals):
        db.add(
            AvailabilityBlock(
                person_id=upload.person_id,
                source=AvailabilitySource.course,
                start_at=start,
                end_at=end,
                status=AvailabilityStatus.active,
                reason="课程",
                source_ref_id=upload.id,
                approved_by=reviewer_id,
                approved_at=now,
            )
        )

    upload.review_status = ReviewStatus.approved
    upload.confirmed_at = now
    db.flush()
    return upload


def reject(db: Session, upload_id: uuid.UUID) -> TimetableUpload:
    upload = _get_upload(db, upload_id)
    upload.review_status = ReviewStatus.rejected
    db.flush()
    return upload


def _supersede_previous(db: Session, current: TimetableUpload) -> None:
    """把同一人同一学期的旧生效课表标记 superseded，并使其课程区间失效。"""
    previous = db.scalars(
        select(TimetableUpload).where(
            TimetableUpload.person_id == current.person_id,
            TimetableUpload.semester_id == current.semester_id,
            TimetableUpload.review_status == ReviewStatus.approved,
            TimetableUpload.id != current.id,
        )
    )
    for up in previous:
        up.review_status = ReviewStatus.superseded
        _expire_blocks_of_upload(db, up.id)


def _expire_blocks_of_upload(db: Session, upload_id: uuid.UUID) -> None:
    blocks = db.scalars(
        select(AvailabilityBlock).where(
            AvailabilityBlock.source == AvailabilitySource.course,
            AvailabilityBlock.source_ref_id == upload_id,
            AvailabilityBlock.status == AvailabilityStatus.active,
        )
    )
    for b in blocks:
        b.status = AvailabilityStatus.expired


def expire_semester_courses(db: Session, semester_id: uuid.UUID) -> int:
    """学期结束：该学期课表逻辑失效，其课程不可值班区间一并失效（方案 4.8）。"""
    uploads = list(
        db.scalars(
            select(TimetableUpload).where(
                TimetableUpload.semester_id == semester_id,
                TimetableUpload.review_status == ReviewStatus.approved,
            )
        )
    )
    for up in uploads:
        up.review_status = ReviewStatus.superseded
        _expire_blocks_of_upload(db, up.id)
    db.flush()
    return len(uploads)


def _get_upload(db: Session, upload_id: uuid.UUID) -> TimetableUpload:
    upload = db.get(TimetableUpload, upload_id)
    if upload is None:
        raise HTTPException(status_code=404, detail="课表上传不存在")
    return upload


def build_free_timetable_excel(db: Session, week: int | None = None) -> bytes:
    import io
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload
    from app.models.enums import PersonStatus, ReviewStatus
    from app.models.person import PersonProfile
    from app.models.timetable import TimetableUpload
    from app.services import semester_service
    from app.timetable.week_parser import parse_weeks

    current_sem = semester_service.get_current_semester(db)
    max_week = current_sem.week_count if current_sem else 20
    active_people = list(
        db.scalars(
            select(PersonProfile)
            .where(PersonProfile.status == PersonStatus.active)
            .order_by(PersonProfile.full_name.asc())
        )
    )

    person_rules: dict[uuid.UUID, list] = {}
    if current_sem:
        stmt = (
            select(TimetableUpload)
            .where(
                TimetableUpload.semester_id == current_sem.id,
                TimetableUpload.review_status == ReviewStatus.approved,
            )
            .options(selectinload(TimetableUpload.course_rules))
        )
        for up in db.scalars(stmt):
            person_rules[up.person_id] = up.course_rules

    PERIOD_BLOCKS = [
        {"label": "1-2 节", "start": 1, "end": 2, "time": "08:00-09:50"},
        {"label": "3-4 节", "start": 3, "end": 4, "time": "10:05-12:10"},
        {"label": "5-6 节", "start": 5, "end": 6, "time": "14:00-15:50"},
        {"label": "7-8 节", "start": 7, "end": 8, "time": "16:05-17:55"},
        {"label": "9-10 节", "start": 9, "end": 10, "time": "19:00-20:50"},
    ]
    WEEKDAYS = [
        (1, "周一"),
        (2, "周二"),
        (3, "周三"),
        (4, "周四"),
        (5, "周五"),
        (6, "周六"),
        (7, "周日"),
    ]

    wb = Workbook()

    # --- Sheet 1: 全员无课表 (按节次) ---
    ws1 = wb.active
    ws1.title = "全员无课表 (按节次)"

    sem_name = current_sem.name if current_sem else "当前学期"
    week_subtitle = f"第 {week} 周" if week else "全学期"
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value = f"全员无课表（可排班人员汇总表） - {sem_name} ({week_subtitle})"
    title_cell.font = Font(name="微软雅黑", size=15, bold=True, color="1F497D")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:H2")
    sub_cell = ws1["A2"]
    sub_cell.value = f"导出日期：{date.today().isoformat()}  |  人员总数：{len(active_people)}人  |  范围：{week_subtitle}  |  说明：列表中为对应时间段无课、可参与班次排班的人员名单"
    sub_cell.font = Font(name="微软雅黑", size=9, italic=True, color="595959")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    headers = ["节次 / 时间", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    ws1.append([])
    ws1.append(headers)
    ws1.row_dimensions[4].height = 28

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx in range(1, 9):
        cell = ws1.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_idx = 5

    def _rule_matches(rule, block_start: int, block_end: int, wd_val: int) -> bool:
        if rule.weekday != wd_val:
            return False
        if rule.period_end < block_start or rule.period_start > block_end:
            return False
        if week is not None and rule.week_expr:
            try:
                spec = parse_weeks(rule.week_expr, max_week)
                if week not in spec.weeks:
                    return False
            except Exception:
                pass
        return True

    for block in PERIOD_BLOCKS:
        row_data = [f"{block['label']}\n({block['time']})"]
        ws1.row_dimensions[row_idx].height = 65

        cell_a = ws1.cell(row=row_idx, column=1, value=row_data[0])
        cell_a.font = Font(name="微软雅黑", size=10, bold=True, color="333333")
        cell_a.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_a.border = thin_border

        for col_i, (wd_val, wd_name) in enumerate(WEEKDAYS, start=2):
            free_names = []
            for person in active_people:
                rules = person_rules.get(person.id, [])
                has_course = any(
                    _rule_matches(r, block["start"], block["end"], wd_val)
                    for r in rules
                )
                if not has_course:
                    free_names.append(person.full_name)

            content = "、".join(free_names) if free_names else "（无）"
            cell_data = ws1.cell(row=row_idx, column=col_i, value=content)
            cell_data.font = Font(name="微软雅黑", size=10, color="262626")
            cell_data.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell_data.border = thin_border

        row_idx += 1

    ws1.column_dimensions["A"].width = 16
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws1.column_dimensions[col_letter].width = 26

    # --- Sheet 2: 人员无课明细 (按个人) ---
    ws2 = wb.create_sheet(title="人员无课明细 (按个人)")
    ws2.append(
        [
            "学号",
            "班级",
            "姓名",
            "手机号",
            "周一无课时段",
            "周二无课时段",
            "周三无课时段",
            "周四无课时段",
            "周五无课时段",
            "周六无课时段",
            "周日无课时段",
        ]
    )
    ws2.row_dimensions[1].height = 28
    header_fill2 = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    for cell in ws2[1]:
        cell.fill = header_fill2
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for person in active_people:
        rules = person_rules.get(person.id, [])
        p_row = [person.student_no, person.class_name, person.full_name, person.phone or ""]
        for wd_val, _ in WEEKDAYS:
            free_blocks = []
            for block in PERIOD_BLOCKS:
                has_course = any(
                    r.weekday == wd_val
                    and not (r.period_end < block["start"] or r.period_start > block["end"])
                    for r in rules
                )
                if not has_course:
                    free_blocks.append(block["label"])
            p_row.append(", ".join(free_blocks) if free_blocks else "全天有课")
        ws2.append(p_row)

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 14
    for c in ["E", "F", "G", "H", "I", "J", "K"]:
        ws2.column_dimensions[c].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
